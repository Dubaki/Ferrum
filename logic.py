import re
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO

class MetalCalculator:
    def __init__(self):
        self.SHEET_LARGE_AREA = 9.0
        self.SHEET_SMALL_AREA = 3.125
        self.PROFILE_STD_LEN = 12000
        self.PROFILE_SMALL_LEN = 6000

    def calculate_sheets(self, parts_data):
        grouped = {}
        for part in parts_data:
            t = part['thickness']
            if t not in grouped: grouped[t] = {'parts': [], 'total_area': 0}
            area = (part['width'] * part['length'] * part['qty']) / 1_000_000
            grouped[t]['parts'].append(part)
            grouped[t]['total_area'] += area

        results = []
        for t, data in grouped.items():
            std_area, std_name = (self.SHEET_SMALL_AREA, "1250x2500") if 0.5 <= t <= 2 else (self.SHEET_LARGE_AREA, "1500x6000")
            count = math.ceil(data['total_area'] / std_area)
            weight = (data['total_area'] * (t / 1000) * 7850) / 1000
            
            results.append({
                'name': f"Лист t={t}мм",
                'summary': f"{count} шт ({std_name})",
                'details': f"Вес: {round(weight, 3)}т, S={round(data['total_area'], 2)}м²",
                'raw_parts': data['parts']
            })
        return results

    def optimize_profile(self, name, parts_list):
        stock_len = self.PROFILE_STD_LEN
        if any(x in name for x in ["15x15", "20x20", "25x25", "30x30", "15x1", "20x2", "35x"]): 
            stock_len = self.PROFILE_SMALL_LEN

        all_pieces = []
        for p in parts_list: all_pieces.extend([p['length']] * p['qty'])
        all_pieces.sort(reverse=True)

        whips = []
        for piece in all_pieces:
            placed = False
            for whip in whips:
                if whip['remaining'] >= piece:
                    whip['remaining'] -= piece
                    whip['parts'].append(piece)
                    placed = True
                    break
            if not placed: whips.append({'remaining': stock_len - piece, 'parts': [piece]})

        total_m = (len(whips) * stock_len) / 1000
        rem_warehouse = sum(w['remaining'] for w in whips if w['remaining'] >= 1000)
        
        return {
            'name': name,
            'summary': f"{len(whips)} хлыстов ({stock_len/1000}м)",
            'details': f"Погонаж: {total_m}м. Склад: {rem_warehouse/1000}м",
            'whips_count': len(whips),
            'stock_len': stock_len,
            'whips_data': whips 
        }

    def generate_excel(self, calculated_data):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заказ металла"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        headers = ["Тип", "Наименование", "Заказ", "Инфо"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        def add_row(cat, name, order, details):
            nonlocal row
            ws.cell(row, 1, cat).border = border
            ws.cell(row, 2, name).border = border
            ws.cell(row, 3, order).border = border
            ws.cell(row, 4, details).border = border
            row += 1

        for item in calculated_data.get('sheets', []):
            add_row("Лист", item['name'], item['summary'], item['details'])

        categories = {
            'square_tubes': 'Квадратная', 'rect_tubes': 'Прямоугольная',
            'es_pipes': 'Труба Э/С', 'angles': 'Уголок', 'rounds': 'Круг'
        }
        for key, label in categories.items():
            for item in calculated_data.get(key, []):
                add_row(label, item['name'], item['summary'], item['details'])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

class SpecParser:
    def __init__(self):
        # Регулярки для очищенного текста (где 'х' заменен на 'x')
        self.re_3_nums = re.compile(r'(\d+)[xх*](\d+)[xх*](\d+[.,]?\d*)', re.I)
        self.re_2_nums = re.compile(r'(\d+)[xх*](\d+[.,]?\d*)', re.I)
        self.re_1_num = re.compile(r'(?:Ø|O|0)?\s*(\d+)', re.I)

    def parse(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        
        data = {'sheets': [], 'rect_tubes': {}, 'square_tubes': {}, 'es_pipes': {}, 'angles': {}, 'rounds': {}}
        
        col_profile = None
        col_length = None
        col_qty = None
        start_row = 0

        # 1. ПОИСК ЗАГОЛОВКОВ (ГЛУБОКИЙ - до 100 строк)
        for r_idx, row in enumerate(rows[:100]):
            row_str = [str(x).lower() for x in row if x]
            if any(k in s for k in ["профиль", "наименование", "сечение"] for s in row_str):
                for c_idx, cell_val in enumerate(row):
                    if not cell_val: continue
                    val = str(cell_val).lower()
                    if "профиль" in val or "наименование" in val: col_profile = c_idx
                    elif ("длина" in val or "мм" in val) and col_length is None: col_length = c_idx
                    elif ("кол" in val or "шт" in val) and col_qty is None: col_qty = c_idx
                
                if col_profile is not None:
                    start_row = r_idx + 1
                    break
        
        # Если не нашли, используем индексы из вашего файла (C, D, E)
        # В Pandas это были бы 2, 3, 4
        if col_profile is None:
            col_qty = 2      # C
            col_profile = 3  # D
            col_length = 4   # E

        # 2. ЧТЕНИЕ ДАННЫХ
        for row in rows[start_row:]:
            if not row or len(row) <= max(col_profile, col_length, col_qty): continue
            
            raw_profile = str(row[col_profile]).strip()
            if not raw_profile or raw_profile.lower() in ["профиль", "none", ""]: continue

            # Чистка цифр
            try:
                len_val = str(row[col_length]).replace(',', '.').replace(u'\xa0', '').strip()
                qty_val = str(row[col_qty]).replace(',', '.').replace(u'\xa0', '').strip()
                if not len_val or len_val == 'None' or not qty_val or qty_val == 'None': continue
                
                length = float(re.findall(r"[\d\.]+", len_val)[0])
                qty = float(re.findall(r"[\d\.]+", qty_val)[0])
            except: continue

            if length <= 0 or qty <= 0: continue

            # --- ЛОГИКА ОПРЕДЕЛЕНИЯ (ПО КЛЮЧЕВЫМ СЛОВАМ) ---
            prof = raw_profile.upper()
            # Важно: нормализуем текст (убираем неразрывные пробелы и меняем кириллицу)
            clean = raw_profile.lower().replace('х', 'x').replace('*', 'x').replace(',', '.').replace(u'\xa0', ' ')

            # 1. ТРУБА ПП
            if "ПП" in prof:
                m = self.re_3_nums.search(clean)
                if m: self._add(data['rect_tubes'], f"Труба ПП {m.group(1)}x{m.group(2)}x{m.group(3)}", length, qty)
                continue

            # 2. ТРУБА ПК
            if "ПК" in prof:
                m3 = self.re_3_nums.search(clean)
                m2 = self.re_2_nums.search(clean)
                if m3: self._add(data['square_tubes'], f"Труба ПК {m3.group(1)}x{m3.group(2)}x{m3.group(3)}", length, qty)
                elif m2: self._add(data['square_tubes'], f"Труба ПК {m2.group(1)}x{m2.group(1)}x{m2.group(2)}", length, qty)
                continue

            # 3. УГОЛОК (L или слово)
            if "УГОЛОК" in prof or prof.startswith("L") or "∟" in prof:
                m3 = self.re_3_nums.search(clean)
                m2 = self.re_2_nums.search(clean)
                if m3: self._add(data['angles'], f"Уголок {m3.group(1)}x{m3.group(2)}x{m3.group(3)}", length, qty)
                elif m2: self._add(data['angles'], f"Уголок {m2.group(1)}x{m2.group(1)}x{m2.group(2)}", length, qty)
                continue

            # 4. ЛИСТ (-, —, Лист)
            if "ЛИСТ" in prof or prof.startswith("-") or "—" in prof or "–" in prof or "PL" in prof:
                m = self.re_2_nums.search(clean)
                if m:
                    data['sheets'].append({'qty': int(qty), 'thickness': float(m.group(1)), 'width': float(m.group(2)), 'length': length})
                continue

            # 5. КРУГ (Ø, Круг)
            if "КРУГ" in prof or (any(x in prof for x in ["O", "0", "Ø"]) and "X" not in prof.replace("X", "x")):
                m = self.re_1_num.search(clean)
                if m: self._add(data['rounds'], f"Круг Ø{m.group(1)}", length, qty)
                continue

            # 6. ТРУБА Э/С
            # Если "Труба" есть, но нет ПП/ПК. Или просто Ø с размерами.
            is_es = ("ТРУБА" in prof and "ПП" not in prof and "ПК" not in prof) or ("Ø" in prof and "X" in prof.replace("X", "x"))
            if is_es:
                m = self.re_2_nums.search(clean)
                if m: self._add(data['es_pipes'], f"Труба Э/С Ø{m.group(1)}x{m.group(2)}", length, qty)
                continue

        return data

    def _add(self, d, name, l, q):
        if name not in d: d[name] = []
        d[name].append({'length': l, 'qty': int(q)})